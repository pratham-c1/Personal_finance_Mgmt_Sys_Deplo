from flask import Blueprint, request, jsonify, send_file, session
from datetime import datetime
import io
from database import execute_query, execute_one

reports_bp = Blueprint('reports', __name__)

def uid(): return session.get('user_id')

@reports_bp.route('/api/reports/monthly', methods=['GET'])
def monthly_report():
    year  = request.args.get('year')
    month = request.args.get('month')
    u     = uid()
    p_income  = [u]
    p_expense = [u]
    y_cond = " AND year_bs=%s"  if year  else ""
    m_cond = " AND month_num=%s" if month else ""
    if year:  p_income.append(year);  p_expense.append(year)
    if month: p_income.append(month); p_expense.append(month)

    income  = execute_query(f"SELECT category, SUM(amount) as total FROM income WHERE user_id=%s{y_cond}{m_cond} GROUP BY category", p_income)
    expense = execute_query(f"SELECT category, SUM(amount) as total FROM expenditure WHERE user_id=%s{y_cond}{m_cond} GROUP BY category", p_expense)
    total_income  = sum(float(r['total']) for r in income)
    total_expense = sum(float(r['total']) for r in expense)
    return jsonify({
        'income_breakdown':  [dict(r) for r in income],
        'expense_breakdown': [dict(r) for r in expense],
        'total_income':      total_income,
        'total_expense':     total_expense,
        'savings':           total_income - total_expense,
    })

@reports_bp.route('/api/reports/export/excel', methods=['GET'])
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        u  = uid()
        wb = openpyxl.Workbook()

        header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        alt_fill    = PatternFill(start_color="f0f4ff", end_color="f0f4ff", fill_type="solid")

        def style_headers(ws, cols):
            for col in range(1, cols+1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

        def auto_width(ws, cols):
            for col in range(1, cols+1):
                ws.column_dimensions[get_column_letter(col)].width = 16

        # Income
        ws = wb.active; ws.title = "Income"
        h = ['ID','Date BS','Date AD','Year','Month','Category','Amount','Remarks']
        ws.append(h); style_headers(ws, len(h))
        for i, r in enumerate(execute_query("SELECT id,date_bs,date_ad,year_bs,month_bs,category,amount,remarks FROM income WHERE user_id=%s ORDER BY date_ad DESC", (u,))):
            ws.append([r['id'],r['date_bs'],str(r['date_ad']),r['year_bs'],r['month_bs'],r['category'],float(r['amount']),r.get('remarks','')])
            if i % 2 == 0:
                for c in range(1, len(h)+1): ws.cell(row=i+2, column=c).fill = alt_fill
        auto_width(ws, len(h))

        # Expenses
        ws2 = wb.create_sheet("Expenses")
        h2  = ['ID','Date BS','Date AD','Category','Particular','Qty','Rate','Amount','Remarks']
        ws2.append(h2); style_headers(ws2, len(h2))
        for i, r in enumerate(execute_query("SELECT id,date_bs,date_ad,category,particular,quantity,rate,amount,remarks FROM expenditure WHERE user_id=%s ORDER BY date_ad DESC", (u,))):
            ws2.append([r['id'],r['date_bs'],str(r['date_ad']),r['category'],r.get('particular',''),float(r.get('quantity',1)),float(r.get('rate',0)),float(r['amount']),r.get('remarks','')])
            if i % 2 == 0:
                for c in range(1, len(h2)+1): ws2.cell(row=i+2, column=c).fill = alt_fill
        auto_width(ws2, len(h2))

        # Net Worth
        ws3 = wb.create_sheet("Net Worth")
        h3  = ['Date BS','Date AD','Bank','Cash','Shares','SSF','Loan Given','Property','Net Worth']
        ws3.append(h3); style_headers(ws3, len(h3))
        for i, r in enumerate(execute_query("SELECT date_bs,date_ad,bank_balance,cash,share_value,ssf,loan_given,property_value,net_worth FROM net_worth WHERE user_id=%s ORDER BY date_ad DESC", (u,))):
            ws3.append([r['date_bs'],str(r['date_ad']),float(r['bank_balance']),float(r['cash']),float(r['share_value']),float(r['ssf']),float(r['loan_given']),float(r['property_value']),float(r['net_worth'])])
            if i % 2 == 0:
                for c in range(1, len(h3)+1): ws3.cell(row=i+2, column=c).fill = alt_fill
        auto_width(ws3, len(h3))

        # Shares
        ws4 = wb.create_sheet("Shares")
        h4  = ['Symbol','Name','Qty','Purchase Price','Current Price','Investment','Current Value','P/L']
        ws4.append(h4); style_headers(ws4, len(h4))
        for r in execute_query("SELECT stock_symbol,stock_name,quantity,purchase_price,current_price,investment,current_value,profit_loss FROM shares WHERE user_id=%s", (u,)):
            ws4.append([r['stock_symbol'],r.get('stock_name',''),r['quantity'],float(r['purchase_price']),float(r['current_price']),float(r['investment']),float(r['current_value']),float(r['profit_loss'])])
        auto_width(ws4, len(h4))

        # Loans
        ws5 = wb.create_sheet("Loans")
        h5  = ['Borrower','Principal','Rate%','Duration(M)','Interest','Total Payable','Status']
        ws5.append(h5); style_headers(ws5, len(h5))
        for r in execute_query("SELECT borrower_name,principal,interest_rate,duration_months,interest_amount,total_payable,status FROM loans WHERE user_id=%s", (u,)):
            ws5.append([r['borrower_name'],float(r['principal']),float(r['interest_rate']),r['duration_months'],float(r['interest_amount']),float(r['total_payable']),r['status']])
        auto_width(ws5, len(h5))

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=f'finance_report_{datetime.now().strftime("%Y%m%d")}.xlsx')
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@reports_bp.route('/api/reports/export/pdf', methods=['GET'])
def export_pdf():
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        u   = uid()
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=1.5*cm, bottomMargin=1.5*cm, leftMargin=1.5*cm, rightMargin=1.5*cm)
        styles  = getSampleStyleSheet()
        story   = []
        hdr_clr = colors.HexColor('#1a1a2e')
        alt_clr = colors.HexColor('#f0f4ff')

        title_s  = ParagraphStyle('t', fontSize=20, fontName='Helvetica-Bold', textColor=hdr_clr, spaceAfter=6, alignment=1)
        h2_s     = ParagraphStyle('h2', fontSize=13, fontName='Helvetica-Bold', textColor=hdr_clr, spaceAfter=4, spaceBefore=14)
        normal_s = ParagraphStyle('n', fontSize=9, fontName='Helvetica', textColor=colors.HexColor('#333'))

        def table_style():
            return TableStyle([
                ('BACKGROUND', (0,0), (-1,0), hdr_clr),
                ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
                ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE',   (0,0), (-1,-1), 9),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, alt_clr]),
                ('GRID',       (0,0), (-1,-1), 0.5, colors.HexColor('#ccc')),
                ('ALIGN',      (1,0), (-1,-1), 'RIGHT'),
                ('TOPPADDING', (0,0), (-1,-1), 5),
                ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ])

        story.append(Paragraph("Personal Finance Report", title_s))
        story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", normal_s))
        story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#e94560')))
        story.append(Spacer(1, 0.3*cm))

        # Summary
        inc = execute_one("SELECT COALESCE(SUM(amount),0) as t FROM income WHERE user_id=%s", (u,))
        exp = execute_one("SELECT COALESCE(SUM(amount),0) as t FROM expenditure WHERE user_id=%s", (u,))
        nw  = execute_one("SELECT net_worth FROM net_worth WHERE user_id=%s ORDER BY date_ad DESC LIMIT 1", (u,))
        ti  = float(inc['t']); te = float(exp['t'])
        summary_data = [
            ['Metric', 'Amount (NPR)'],
            ['Total Income',  f"Rs. {ti:,.2f}"],
            ['Total Expense', f"Rs. {te:,.2f}"],
            ['Savings',       f"Rs. {ti-te:,.2f}"],
            ['Net Worth',     f"Rs. {float(nw['net_worth']):,.2f}" if nw else 'N/A'],
        ]
        story.append(Paragraph("Financial Summary", h2_s))
        t = Table(summary_data, colWidths=[8*cm, 8*cm])
        t.setStyle(table_style()); story.append(t); story.append(Spacer(1, 0.5*cm))

        # Income breakdown
        inc_cat = execute_query("SELECT category, SUM(amount) as total FROM income WHERE user_id=%s GROUP BY category ORDER BY total DESC", (u,))
        if inc_cat:
            story.append(Paragraph("Income Breakdown", h2_s))
            it = Table([['Category','Amount (NPR)']] + [[r['category'], f"Rs. {float(r['total']):,.2f}"] for r in inc_cat], colWidths=[8*cm, 8*cm])
            it.setStyle(table_style()); story.append(it); story.append(Spacer(1, 0.5*cm))

        # Expense breakdown
        exp_cat = execute_query("SELECT category, SUM(amount) as total FROM expenditure WHERE user_id=%s GROUP BY category ORDER BY total DESC", (u,))
        if exp_cat:
            story.append(Paragraph("Expense Breakdown", h2_s))
            et = Table([['Category','Amount (NPR)']] + [[r['category'], f"Rs. {float(r['total']):,.2f}"] for r in exp_cat], colWidths=[8*cm, 8*cm])
            et.setStyle(table_style()); story.append(et)

        # Shares
        shares = execute_query("SELECT stock_symbol,quantity,purchase_price,current_price,investment,current_value,profit_loss FROM shares WHERE user_id=%s", (u,))
        if shares:
            story.append(Spacer(1, 0.5*cm))
            story.append(Paragraph("Share Portfolio", h2_s))
            sdata = [['Symbol','Qty','Buy','Curr','Invested','Value','P/L']]
            for r in shares:
                sdata.append([r['stock_symbol'],r['quantity'],f"Rs.{float(r['purchase_price']):,.0f}",f"Rs.{float(r['current_price']):,.0f}",f"Rs.{float(r['investment']):,.0f}",f"Rs.{float(r['current_value']):,.0f}",f"Rs.{float(r['profit_loss']):,.0f}"])
            st = Table(sdata, colWidths=[2.3*cm]*7)
            st.setStyle(table_style()); story.append(st)

        doc.build(story); buf.seek(0)
        return send_file(buf, mimetype='application/pdf', as_attachment=True,
                         download_name=f'finance_report_{datetime.now().strftime("%Y%m%d")}.pdf')
    except Exception as e:
        return jsonify({'error': str(e)}), 500
